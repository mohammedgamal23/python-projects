import openpyxl as xl
from openpyxl.chart import BarChart, Reference

#prompt the user for a spreedsheet file
file1 = input("Enter a filename: ")

#load the file
wb = xl.load_workbook('file1')
sheet = wb['Sheet1']

#access data which is in third column to manipulate
for row in range(2,sheet.max_row+1):
    #save every piece of data in a cell
    cell = sheet.cell(row,3)
    
    #manipulate data
    corrected_price = cell.value * 0.9
    
    #add it the the forth column
    corrected_price_cell = sheet.cell(row,4)
    corrected_price_cell.value = corrected_price

#get a reference of the values to make a bar chart
values = Reference(sheet,min_row=2,max_row=sheet.max_row,min_col=4, max_col=4)
chart = BarChart()
chart.add_data(values)

#start the chart from E2 cell
sheet.add_chart(chart, 'E2')

#save the new manipulated file in a new file
wb.save('file3.xlsx')
